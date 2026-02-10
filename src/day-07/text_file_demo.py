import csv
from openpyxl import load_workbook,workbook
from openpyxl import Workbook

file=open("sample.txt","w")
x=file.write("hello, writing the python...")
file.close()

file1 = open("sample.txt","r")
y=file1.read()
print(y)

with open("sample.txt","r") as file:
    reads = file.read()
    print(reads)


try:
    with open("sample1.txt","r") as file:
        print(file.read())
except FileNotFoundError:
    print("ERROR: File Not Found, try with another file")



with open("Sample_Data.csv", "r") as file:
    reader = csv.reader(file)
    for row in reader:
        print(row)



data = load_workbook("student.xlsx")
datas = data.active

for row in range(0,datas.max_row):
    for col in datas.iter_cols(1,datas.max_column):
        print(col[row].value)



# Writing the Own CSV file
with open("day6.csv","w",newline='') as file:
    write = csv.writer(file)
    write.writerow(['Name','Age','Class','Year'])
    write.writerow(['Abhi','21','B.E','2005'])
    write.writerow(['Avii','20','B.Tech','2004'])
    write.writerow(['Akhi','16','10','2010'])
    write.writerow(['mann','19','11','2007'])
with open('day6.csv','r') as reads:
        readerx = reads.read()
        print(readerx)


# Creating out OWN EXCEL file
wb = Workbook()
sheet = wb.active

sheet["A1"] = "Name"
sheet["A2"] = "Abhi"
sheet["A3"] = "Avi"
sheet["A4"] = "Akhi"

sheet["B1"] = "Age"
sheet["B2"] = "21"
sheet["B3"] = "20"
sheet["B4"] = "30"

sheet["C1"] = "Class"
sheet["C2"] = "12"
sheet["C3"] = "11"
sheet["C4"] = "10"

sheet["D1"] = "Year"
sheet["D2"] = "2026"
sheet["D3"] = "2025"
sheet["D4"] = "2050"

wb.save("details.xlsx")